from table import Table
from tkinter import Label, Entry, Button, OptionMenu, Checkbutton, IntVar, StringVar, CENTER
from choosing_file import choosing_file
from preparing_data import inform
import json
import sys
import os


def update_summary_data(button, center, top_frame):
    names_to_remember.extend(table.summary_view())
    button.destroy()

    _summary_entries(center, top_frame)


def _summary_entries(center, top_frame):

    table.destroy_buttons()

    date = Label(top_frame, text="Date:")
    date.grid(row=0, column=1)

    entry_date = Entry(top_frame, bg="white", fg="black", width=10)
    entry_date.grid(row=0, column=2)
    entry_date.insert(0, "YYYY-MM-DD")

    var = StringVar()
    var.set("Cash/Card")
    cash_card = OptionMenu(top_frame, var, "Cash", "Card")
    cash_card.grid(row=0, column=3)

    empty = Label(top_frame, text="         ")
    empty.grid(row=0, column=5)

    button_export = Button(top_frame, text="Export",
                        command=lambda: exporting_data(entry_date, var, var_check, sheet, sheet_col,
                                        split_sum, split_sheet, split_sheet_col, center, top_frame, entry_path=None))
    button_export.grid(row=0, column=6)

    if not google_sheet:
        entry = _excel_gui(top_frame)

    var_check = IntVar()
    path_remember = Checkbutton(top_frame, text="Remember", variable=var_check)
    path_remember.grid(row=1, column=4)

    sheet_label = Label(top_frame, text="Sheet and column:")
    sheet_label.grid(row=2, column=1)

    sheet = Entry(top_frame, bg="white", fg="black", width=10)
    sheet.grid(row=2, column=2)
    sheet.insert(0, "Sheet")

    sheet_col = Entry(top_frame, bg="white", fg="black", width=3, justify=CENTER)
    sheet_col.grid(row=2, column=3)
    sheet_col.insert(0, "B")

    split_sum = check_split()

    if split_sum:
        split_sheet, split_sheet_col = _split_sum_gui(top_frame)

    try:
        with open("saves.json", 'r') as f:
            score = json.load(f)

            var_check.set(score[0])
            if not google_sheet:
                entry.insert(0, str(score[1]))

            sheet.delete(0, 'end')
            sheet.insert(0, str(score[2]))

            sheet_col.delete(0, 'end')
            sheet_col.insert(0, str(score[3]))

            if split_sum:

                split_sheet.delete(0, 'end')
                split_sheet.insert(0, str(score[4]))

                split_sheet_col.delete(0, 'end')
                split_sheet_col.insert(0, str(score[5]))

            else:
                split_sheet = str(score[4])
                split_sheet_col = str(score[5])

    except Exception as e:
        split_sheet = None
        split_sheet_col = None

        inform(f"Error: {e}.")


def _excel_gui(top_frame):
    api_or = Label(top_frame, text="File location:")
    api_or.grid(row=1, column=1)

    entry = Entry(top_frame, bg="white", fg="black", width=10)
    entry.grid(row=1, column=2)

    choose_file = Button(top_frame, text="Choose a file", command=lambda: choosing_file(entry, "xlsx"))
    choose_file.grid(row=1, column=3)
    entry_path = entry.get()

    return entry


def _split_sum_gui(top_frame):
    split_sheet_label = Label(top_frame, text="Sheet and column of split data:")
    split_sheet_label.grid(row=3, column=1)

    split_sheet = Entry(top_frame, bg="white", fg="black", width=10)
    split_sheet.grid(row=3, column=2)

    split_sheet.insert(0, "Sheet")
    split_sheet_col = Entry(top_frame, bg="white", fg="black", width=3, justify=CENTER)
    split_sheet_col.grid(row=3, column=3)

    split_sheet_col.insert(0, "B")

    return split_sheet, split_sheet_col


def exporting_data(entry_date, var, var_check, sheet, sheet_col, split_sum, split_sheet, split_sheet_col, center, top_frame, entry_path):

    if split_sum:
        split_sheet = split_sheet.get()
        split_sheet_col = split_sheet_col.get()

    if var_check.get():
        with open("saves.json", 'w') as f:
            save = [var_check.get(), entry_path, sheet.get(), sheet_col.get(), split_sheet, split_sheet_col]
            json.dump(save, f, indent=2)

    if any(item == "" for item in [entry_date.get(), entry_path, sheet.get(), sheet_col.get()]) or var.get() == "Cash/Card":
        return

    sub_type = "Cash" if var.get() == "Cash" else "Regular"

    import pandas as pd

    data_main = [[entry_date.get(), "Expense", "Accounts", var.get(), sub_type, row[1].cget("text"), row[2].get(), _cost_convert(row[3].cget("text"))] for row in table.widgets]
    df_1 = pd.DataFrame(data_main)

    if google_sheet:

        from google_sheets_export import google_sheets_export

        google_sheets_export(df_1, sheet.get(), sheet_col.get())

        if split_sum:
            df_2 = _split_sheet_export(entry_date, split_sum, var, pd)
            google_sheets_export(df_2, split_sheet, split_sheet_col)
    else:

        from openpyxl.utils import column_index_from_string

        column_index = (column_index_from_string(sheet_col.get()) - 1)
        with pd.ExcelWriter(fr'{entry_path}', mode='a', if_sheet_exists='overlay') as writer:
            df_1.to_excel(writer, sheet_name=sheet.get(), startrow=find_empty_row(entry_path, sheet.get()),
                          startcol=column_index, index=False, header=False, engine='openpyxl')

        if split_sum:
            df_2 = _split_sheet_export(entry_date, split_sum, var, pd)

            column_index = (column_index_from_string(split_sheet_col) - 1)
            with pd.ExcelWriter(fr'{entry_path}', mode='a', if_sheet_exists='overlay') as writer:
                df_2.to_excel(writer, sheet_name=sheet.get(), startrow=find_empty_row(entry_path, split_sheet),
                              startcol=column_index, index=False, header=False, engine='openpyxl')

    table.destroy_info()

    for widget in top_frame.grid_slaves():
        widget.destroy()

    save_gui(center, top_frame)


def _split_sheet_export(entry_date, split_sum, var, pd):
    split_sum = _cost_convert(split_sum)
    data_split = [entry_date.get(), "L-->M", var.get(), split_sum]
    df_2 = pd.DataFrame(data_split).transpose()
    return df_2


def save_gui(center, top_frame):
    var_delete = IntVar()
    delete_data = Checkbutton(top_frame, text="Delete data file", variable=var_delete)
    delete_data.grid(row=0, columnspan=2)

    save_quit = Button(top_frame, text="Save & quit", command=lambda: save_and_quit(remember_saved, var_delete.get()))
    save_quit.grid(row=1, column=1)

    table_name = Label(center, text="Name")
    table_name.grid(row=0, column=0)

    table_category = Label(center, text="Category")
    table_category.grid(row=0, column=1)

    table_save = Label(center, text="Save")
    table_save.grid(row=0, column=2)

    remember_saved = []

    for count, name in enumerate(names_to_remember):

        info_name = Label(center, text=name[0])
        info_name.grid(row=count + 1, column=0)

        info_subcat = Label(center, text=name[1])
        info_subcat.grid(row=count + 1, column=1)

        var_save = IntVar()
        to_remember = Checkbutton(center, text=" ", variable=var_save)
        to_remember.grid(row=count + 1, column=2)

        remember_saved.append([info_name, info_subcat, var_save])


def check_split():
    total = 0
    for row in table.widgets:
        if not row[4].cget("text") == "         ":
            total += float(row[4].cget("text"))

    return total


def save_and_quit(remember_saved, var_delete):
    _save_names(remember_saved)

    if var_delete:
        _delete_json()

    sys.exit()


def _delete_json():
    os.remove(json_name + ".json")


def _save_names(save):
    with open("options.json", 'r') as f:
        options = json.load(f)

    for data in save:
        if data[2].get():
            options[data[1].cget("text")].append(data[0].cget("text").lower())

    with open("options.json", 'w') as g:
        json.dump(options, g, indent=2, ensure_ascii=False)


def _cost_convert(number):
    number = str(number).replace(".", ",")

    return number


def find_empty_row(document_name, sheet):

    from openpyxl import load_workbook
    workbook = load_workbook(document_name)
    worksheet = workbook[sheet]

    row = worksheet.max_row

    return row


def main_table(data, drop_down_list, center, top_frame, google, json):

    global table, names_to_remember, google_sheet, json_name
    google_sheet, json_name = google, json

    names_to_remember = []

    table = Table(center, data, drop_down_list)
    button = Button(top_frame, text="Summary view", command=lambda: update_summary_data(button, center, top_frame))
    button.grid()
